"""
================================================================
 PART 3: PIPELINE
 Reads assignment_cases.csv, runs each report through the
 parser + classifier, and writes assignment_cases_updated.csv
 + a formatted Excel workbook.
================================================================
"""

import csv
import sys
import os

# Add the directory containing our modules to path
sys.path.insert(0, '/home/claude')

from report_parser    import parse_report
from classifier_engine import classify


# ── Column order for output CSV ──────────────────────────────
INPUT_COLS  = ["Case_ID", "Type", "Report"]
OUTPUT_COLS = [
    "Your_Risk_Category",
    "Your_Nodule_Type",
    "Your_Nodule_Size",
    "Your_Nodule_Count",
    "Your_Category_or_Guideline",
    "Your_Management_Recommendation",
    "Your_Reasoning",
]
ALL_COLS = INPUT_COLS + OUTPUT_COLS


def format_risk_category(result, case_type):
    if case_type == "Lung-RADS":
        return "N/A (Lung-RADS)"
    return result.get("risk_category", "")


def format_nodule_type(result, case_type):
    if case_type == "Lung-RADS":
        return result.get("nodule_type", "Solid")
    return result.get("nodule_type", "")


def format_nodule_size(result, case_type):
    return result.get("nodule_size", "N/A")


def format_nodule_count(result, case_type):
    if case_type == "Lung-RADS":
        return "N/A (Lung-RADS)"
    return result.get("nodule_count", "")


def format_category(result, case_type):
    if case_type == "Lung-RADS":
        return result.get("category", "")
    return result.get("category", "")


def format_recommendation(result):
    return result.get("recommendation", "")


def format_reasoning(result, features):
    """
    Builds a full human-readable reasoning string combining:
    - Extracted features from the parser
    - Classification reasoning from the engine
    """
    lines = []

    # Step 1: Eligibility / guideline selection
    if features.get("is_screening"):
        lines.append("GUIDELINE SELECTION: Lung-RADS v2022 — this is a lung cancer screening LDCT.")
    else:
        lines.append("GUIDELINE SELECTION: Fleischner 2017 — incidental nodule on diagnostic CT.")

    # Step 2: Exclusion / eligibility (Fleischner only)
    if not features.get("is_screening"):
        excl = []
        if features.get("age", 99) < 35:    excl.append("Age < 35")
        if features.get("known_cancer"):     excl.append("Known primary cancer")
        if features.get("immunocompromised"):excl.append("Immunocompromised")
        if excl:
            lines.append(f"EXCLUSION TRIGGERED: {'; '.join(excl)}")
        else:
            lines.append(f"ELIGIBILITY: Age {features.get('age')} ≥ 35 | "
                         "No known cancer | Not immunocompromised | Not screening → Fleischner applies.")

        # Step 3: Risk factors
        rf = result.get("risk_factors", [])
        if rf:
            lines.append(f"RISK STRATIFICATION: {result.get('risk_category', '')} — "
                         f"{'; '.join(rf)}.")
        else:
            lines.append("RISK STRATIFICATION: Low Risk — no high-risk features identified.")

    # Step 4: Nodule characterization
    ntype  = features.get("nodule_type", "Solid")
    size   = features.get("nodule_size_mm", 0)
    long_  = features.get("nodule_long_mm", size)
    short_ = features.get("nodule_short_mm", size)
    count  = features.get("nodule_count", "single")
    status = features.get("nodule_status", "baseline")

    if long_ != short_:
        size_str = f"({long_} + {short_}) / 2 = {round((long_+short_)/2,1)} mm"
    else:
        size_str = f"{size} mm"

    if features.get("is_screening"):
        lines.append(f"NODULE CHARACTERIZATION: Type={ntype}, Size={size_str}, "
                     f"Status={status} (baseline/new/growing/stable).")
    else:
        lines.append(f"NODULE CHARACTERIZATION: Type={ntype}, Size={size_str}, "
                     f"Count={count}.")

    # Step 5: Special cases
    if features.get("incomplete_evaluation"):
        lines.append("SPECIAL CASE: Incomplete evaluation — nodule cannot be fully "
                     "characterized from this study. Dedicated CT chest required.")

    # Step 6: The classification rationale from the engine
    lines.append(f"CLASSIFICATION RATIONALE: {result.get('reasoning', '')}")

    # Step 7: Lung-RADS specific extras
    if features.get("is_screening"):
        s_mods = result.get("s_modifier_findings")
        if s_mods:
            lines.append(f"S MODIFIER: {'; '.join(s_mods)}.")
        x_feats = result.get("4x_features")
        if x_feats:
            lines.append(f"4X UPGRADE TRIGGERED BY: {'; '.join(x_feats)}.")
        lines.append(f"SCREEN RESULT: {result.get('screen_result', '')}.")

    return " | ".join(lines)


def process_csv(input_path, output_csv_path):
    """
    Main pipeline:
    1. Read each row from input CSV
    2. Parse the report text → feature dict
    3. Classify → result dict
    4. Format output columns
    5. Write to output CSV
    Returns list of (row_dict, features, result) for Excel generation.
    """
    processed_rows = []

    with open(input_path, newline='', encoding='utf-8') as f_in:
        reader = csv.DictReader(f_in)

        with open(output_csv_path, 'w', newline='', encoding='utf-8') as f_out:
            writer = csv.DictWriter(f_out, fieldnames=ALL_COLS)
            writer.writeheader()

            for row in reader:
                case_id   = row["Case_ID"]
                case_type = row["Type"]        # "Fleischner" or "Lung-RADS"
                report    = row["Report"]

                # ── Step 1: Extract features from free text ──
                features = parse_report(report, case_type)

                # ── Step 2: Classify ─────────────────────────
                result = classify(features)

                # ── Step 3: Format output columns ────────────
                out_row = {
                    "Case_ID":  case_id,
                    "Type":     case_type,
                    "Report":   report,
                    "Your_Risk_Category":           format_risk_category(result, case_type),
                    "Your_Nodule_Type":             format_nodule_type(result, case_type),
                    "Your_Nodule_Size":             format_nodule_size(result, case_type),
                    "Your_Nodule_Count":            format_nodule_count(result, case_type),
                    "Your_Category_or_Guideline":   format_category(result, case_type),
                    "Your_Management_Recommendation": format_recommendation(result),
                    "Your_Reasoning":               format_reasoning(result, features),
                }
                writer.writerow(out_row)
                processed_rows.append((out_row, features, result))
                print(f"  Processed {case_id}: {result.get('category', result.get('lung_rads_category', '?'))}")

    return processed_rows


def build_excel(processed_rows, output_xlsx_path):
    """
    Generates a formatted Excel file from the processed rows.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Classification Results"

    # Styles
    def border():
        s = Side(style="thin", color="AAAAAA")
        return Border(left=s, right=s, top=s, bottom=s)

    def hdr(cell, bg="1F3864", fg="FFFFFF"):
        cell.font      = Font(name="Arial", bold=True, color=fg, size=10)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border()

    def body(cell, bg="FFFFFF", bold=False, center=False):
        cell.font      = Font(name="Arial", size=9, bold=bold)
        cell.fill      = PatternFill("solid", fgColor=bg)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="top", wrap_text=True
        )
        cell.border = border()

    # Headers
    display_headers = [
        "Case ID", "Type",
        "Risk Category", "Nodule Type", "Nodule Size",
        "Nodule Count", "Category / Guideline",
        "Management Recommendation", "Reasoning"
    ]
    for col, h in enumerate(display_headers, 1):
        hdr(ws.cell(row=1, column=col, value=h))

    # Row color coding
    ROW_COLORS = {
        "Fleischner": {"low": "EBF5FB", "high": "D6EAF8"},
        "Lung-RADS":  {"low": "E9F7EF", "high": "D5F5E3"},
    }
    WARN_COLOR = "FADBD8"  # Pink for urgent/4B cases

    for row_idx, (out_row, features, result) in enumerate(processed_rows, 2):
        case_type = out_row["Type"]
        cat       = str(result.get("category", result.get("lung_rads_category", "")))
        is_urgent = any(x in cat for x in ["4B", "4X", "> 8 mm", "HIGHLY SUSPICIOUS"])
        is_alt    = row_idx % 2 == 0

        if is_urgent:
            bg = WARN_COLOR
        elif case_type == "Fleischner":
            bg = ROW_COLORS["Fleischner"]["low" if is_alt else "high"]
        else:
            bg = ROW_COLORS["Lung-RADS"]["low" if is_alt else "high"]

        values = [
            out_row["Case_ID"],
            out_row["Type"],
            out_row["Your_Risk_Category"],
            out_row["Your_Nodule_Type"],
            out_row["Your_Nodule_Size"],
            out_row["Your_Nodule_Count"],
            out_row["Your_Category_or_Guideline"],
            out_row["Your_Management_Recommendation"],
            out_row["Your_Reasoning"],
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            body(cell, bg=bg, bold=(col <= 2), center=(col <= 6))

    # Column widths
    widths = [9, 12, 18, 14, 20, 14, 38, 45, 95]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row heights
    ws.row_dimensions[1].height = 30
    for r in range(2, len(processed_rows) + 2):
        ws.row_dimensions[r].height = 100

    ws.freeze_panes = "C2"

    wb.save(output_xlsx_path)
    print(f"\n  Excel saved: {output_xlsx_path}")


# ================================================================
#  MAIN
# ================================================================

if __name__ == "__main__":
    INPUT_CSV   = "assignment_cases_updated.csv"
    OUTPUT_CSV  = "output_assignment_cases_updated.csv"
    OUTPUT_XLSX = "assignment_cases_updated.xlsx"

    print("\n" + "="*60)
    print("  PULMONARY NODULE CLASSIFICATION PIPELINE")
    print("="*60)
    print(f"\n  Input : {INPUT_CSV}")
    print(f"  Output CSV : {OUTPUT_CSV}")
    print(f"  Output XLSX: {OUTPUT_XLSX}\n")

    print("  Processing cases...")
    rows = process_csv(INPUT_CSV, OUTPUT_CSV)

    print("\n  Building Excel workbook...")
    build_excel(rows, OUTPUT_XLSX)

    print(f"\n  Done. {len(rows)} cases processed.")
    print("="*60)
